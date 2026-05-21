# -*- coding: utf-8 -*-
"""
Экспорт данных анализа в Excel (XLSX).

Использует openpyxl для создания файла с несколькими листами:
- Метаданные
- Временные ряды
- Спектры
- Результаты анализа
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any
import numpy as np

from ..dal.logger import get_logger

logger = get_logger("ExcelExporter")

# Проверяем наличие openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl не установлен. Excel-экспорт недоступен.")


class ExcelExporter:
    """Экспорт результатов анализа в Excel-файл."""

    def __init__(self, parser, sensor_id: int):
        """
        Инициализация экспортера.

        Args:
            parser: Парсер с данными (MultiSensorRD2Parser).
            sensor_id: ID датчика для экспорта.
        """
        self.parser = parser
        self.sensor_id = sensor_id
        self.data = parser.get_sensor_data(sensor_id) if parser else None

        # Стиль заголовков (чёрно-белая тема) — лениво, только при наличии openpyxl
        if HAS_OPENPYXL:
            self._header_font = Font(bold=True, color="FFFFFF", size=11)
            self._header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            self._header_alignment = Alignment(horizontal="center", vertical="center")
            self._data_font = Font(color="000000", size=10)
            self._thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

    def export(self, file_path: Path) -> bool:
        """
        Экспортировать данные в Excel.

        Args:
            file_path: Путь для сохранения файла.

        Returns:
            True если экспорт успешен.
        """
        if not HAS_OPENPYXL:
            logger.error("openpyxl не установлен. Установите: pip install openpyxl")
            return False

        if self.data is None:
            logger.error("Нет данных для экспорта (sensor_id=%d)", self.sensor_id)
            return False

        try:
            file_path = Path(file_path)
            logger.info("Экспорт Excel: %s", file_path)

            wb = Workbook()

            # Создаём листы
            self._create_metadata_sheet(wb)
            self._create_time_series_sheet(wb)
            self._create_spectrum_sheet(wb)
            self._create_analysis_sheet(wb)

            # Удаляем дефолтный лист
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            wb.save(str(file_path))
            logger.info("Excel экспорт завершён: %s", file_path)
            return True

        except Exception as e:
            logger.error("Ошибка экспорта Excel: %s", e, exc_info=True)
            return False

    def _create_metadata_sheet(self, wb: 'Workbook'):
        """Создать лист с метаданными турбины."""
        ws = wb.create_sheet("Метаданные")
        metrics = self.parser.get_turbine_metrics()

        # Заголовок
        ws['A1'] = 'Параметр'
        ws['B1'] = 'Значение'
        ws['C1'] = 'Единица'
        self._style_header_row(ws, 1)

        rows = [
            ['Мощность', f"{metrics.get('power_kw', 0):.1f}", 'кВт'],
            ['Частота вращения', f"{metrics.get('generator_speed_rpm', 0):.1f}", 'об/мин'],
            ['Скорость ветра', f"{metrics.get('wind_speed_ms', 0):.1f}", 'м/с'],
            ['Накопленная выработка', f"{metrics.get('cumulative_power_kwh', 0):.1f}", 'кВт·ч'],
        ]
        for idx, row in enumerate(rows, start=2):
            ws.append(row)
            self._style_data_row(ws, idx)

        # Ширина колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12

    def _create_time_series_sheet(self, wb: 'Workbook'):
        """Создать лист с временными рядами."""
        ws = wb.create_sheet("Временные ряды")

        ws['A1'] = 'Тип сигнала'
        ws['B1'] = 'Время (с)'
        ws['C1'] = 'Значение'
        self._style_header_row(ws, 1)

        row_idx = 2
        for signal_type, signal_key, time_key in [
            ('НЧ (Ускорение)', 'acceleration', 'acceleration_time'),
            ('ВЧ (Скорость)', 'velocity', 'velocity_time'),
            ('ВЧ(ф) (ВЧ фильтр)', 'high_freq', 'high_freq_time'),
        ]:
            signal = self.data.get(signal_key)
            times = self.data.get(time_key)
            if signal is not None and times is not None:
                for t, val in zip(times, signal):
                    ws.append([signal_type, f"{t:.6f}", f"{val:.6f}"])
                    self._style_data_row(ws, row_idx)
                    row_idx += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_spectrum_sheet(self, wb: 'Workbook'):
        """Создать лист со спектрами."""
        from ..utils.vibration_analysis import VibrationAnalyzer
        analyzer = VibrationAnalyzer()

        ws = wb.create_sheet("Спектры")

        ws['A1'] = 'Тип сигнала'
        ws['B1'] = 'Частота (Гц)'
        ws['C1'] = 'Амплитуда'
        self._style_header_row(ws, 1)

        row_idx = 2
        for signal_type, signal_key, fs_key in [
            ('НЧ (Ускорение)', 'acceleration', 'acceleration_fs'),
            ('ВЧ (Скорость)', 'velocity', 'velocity_fs'),
            ('ВЧ(ф) (ВЧ фильтр)', 'high_freq', 'high_freq_fs'),
        ]:
            signal = self.data.get(signal_key)
            fs = self.data.get(fs_key)
            if signal is not None and fs:
                freqs, amps = analyzer.calculate_spectrum(np.array(signal), fs)
                for f, a in zip(freqs, amps):
                    ws.append([signal_type, f"{f:.3f}", f"{a:.6f}"])
                    self._style_data_row(ws, row_idx)
                    row_idx += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_analysis_sheet(self, wb: 'Workbook'):
        """Создать лист с результатами анализа."""
        from ..utils.vibration_analysis import VibrationAnalyzer
        analyzer = VibrationAnalyzer()

        ws = wb.create_sheet("Результаты анализа")

        ws['A1'] = 'Тип сигнала'
        ws['B1'] = 'RMS'
        ws['C1'] = 'Единица'
        ws['D1'] = 'Зона ISO 10816'
        self._style_header_row(ws, 1)

        rows = []
        if self.data.get('acceleration') is not None:
            rms = np.sqrt(np.mean(np.array(self.data['acceleration']) ** 2))
            zone = analyzer.determine_zone_acc(rms)
            rows.append(['НЧ (Ускорение)', f"{rms:.6f}", 'м/с²', zone])

        if self.data.get('velocity') is not None:
            rms = np.sqrt(np.mean(np.array(self.data['velocity']) ** 2))
            zone = analyzer.determine_zone_vel(rms)
            rows.append(['ВЧ (Скорость)', f"{rms:.6f}", 'мм/с', zone])

        if self.data.get('high_freq') is not None:
            rms = np.sqrt(np.mean(np.array(self.data['high_freq']) ** 2))
            zone = analyzer.determine_zone_acc(rms)
            rows.append(['ВЧ(ф) (ВЧ фильтр)', f"{rms:.6f}", 'м/с²', zone])

        for idx, row in enumerate(rows, start=2):
            ws.append(row)
            self._style_data_row(ws, idx)

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18

    def _style_header_row(self, ws, row: int):
        """Применить стиль к строке заголовков."""
        if not HAS_OPENPYXL:
            return
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_alignment
            cell.border = self._thin_border

    def _style_data_row(self, ws, row: int):
        """Применить стиль к строке данных."""
        if not HAS_OPENPYXL:
            return
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self._data_font
            cell.border = self._thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
