# -*- coding: utf-8 -*-
"""
Unit-тесты для модулей экспорта.
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, MagicMock
import numpy as np

from smp12c_vibrodiag.exporters.csv_exporter import CSVExporter
from smp12c_vibrodiag.exporters.excel_exporter import ExcelExporter


class MockParser:
    """Фейковый парсер для тестов."""

    def __init__(self):
        self._metrics = {
            'power_kw': 2500.5,
            'generator_speed_rpm': 1500.0,
            'wind_speed_ms': 12.5,
            'cumulative_power_kwh': 15000000.0,
        }
        self._sensor_data = {
            1: {
                'acceleration': np.random.randn(100) * 0.5,
                'acceleration_time': np.linspace(0, 10, 100),
                'acceleration_fs': 100.0,
                'velocity': np.random.randn(100) * 2.0,
                'velocity_time': np.linspace(0, 10, 100),
                'velocity_fs': 100.0,
                'high_freq': np.random.randn(100) * 0.3,
                'high_freq_time': np.linspace(0, 10, 100),
                'high_freq_fs': 100.0,
            }
        }

    def get_turbine_metrics(self):
        return self._metrics

    def get_sensor_data(self, sensor_id):
        return self._sensor_data.get(sensor_id)

    def get_available_sensors(self):
        return list(self._sensor_data.keys())


class TestCSVExporter:
    """Тесты CSV экспортера."""

    def setup_method(self):
        """Инициализация перед каждым тестом."""
        self.parser = MockParser()
        self.exporter = CSVExporter(self.parser, sensor_id=1)

    def test_export_creates_file(self, temp_dir):
        """Экспорт создаёт файл."""
        file_path = temp_dir / "test_export.csv"
        success = self.exporter.export(file_path)
        
        assert success is True
        assert file_path.exists()
        assert file_path.stat().st_size > 0

    def test_export_file_content(self, temp_dir):
        """Проверка содержимого CSV."""
        file_path = temp_dir / "test_export.csv"
        self.exporter.export(file_path)
        
        content = file_path.read_text(encoding='utf-8-sig')
        assert '=== МЕТАДАННЫЕ ТУРБИНЫ ===' in content
        assert '=== ВРЕМЕННЫЕ РЯДЫ ===' in content
        assert '=== СПЕКТРЫ ===' in content
        assert '=== РЕЗУЛЬТАТЫ АНАЛИЗА' in content
        assert 'Мощность' in content
        assert '2500.5' in content

    def test_export_no_data(self, temp_dir):
        """Экспорт без данных."""
        empty_parser = Mock()
        empty_parser.get_sensor_data = Mock(return_value=None)
        exporter = CSVExporter(empty_parser, sensor_id=1)
        
        file_path = temp_dir / "empty.csv"
        success = exporter.export(file_path)
        
        assert success is False

    def test_export_nonexistent_path(self):
        """Экспорт в несуществующую директорию."""
        file_path = Path("/nonexistent/dir/test.csv")
        success = self.exporter.export(file_path)
        assert success is False


class TestExcelExporter:
    """Тесты Excel экспортера."""

    def setup_method(self):
        """Инициализация перед каждым тестом."""
        self.parser = MockParser()
        self.exporter = ExcelExporter(self.parser, sensor_id=1)

    def test_export_creates_file(self, temp_dir):
        """Экспорт создаёт файл XLSX."""
        file_path = temp_dir / "test_export.xlsx"
        success = self.exporter.export(file_path)
        
        assert success is True
        assert file_path.exists()
        assert file_path.stat().st_size > 0

    def test_export_sheets(self, temp_dir):
        """Проверка листов Excel."""
        from openpyxl import load_workbook
        
        file_path = temp_dir / "test_export.xlsx"
        self.exporter.export(file_path)
        
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        
        assert 'Метаданные' in sheet_names
        assert 'Временные ряды' in sheet_names
        assert 'Спектры' in sheet_names
        assert 'Результаты анализа' in sheet_names

    def test_export_no_data(self, temp_dir):
        """Экспорт без данных."""
        empty_parser = Mock()
        empty_parser.get_sensor_data = Mock(return_value=None)
        exporter = ExcelExporter(empty_parser, sensor_id=1)
        
        file_path = temp_dir / "empty.xlsx"
        success = exporter.export(file_path)
        
        assert success is False
